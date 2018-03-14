from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 43

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

ws1 = wb.create_sheet('Mysheet')

# print all sheetnames
print(wb.sheetnames)

# loop over sheets

for sheet in wb:
    print(sheet.title)


# Save the file
wb.save("sample.xlsx")